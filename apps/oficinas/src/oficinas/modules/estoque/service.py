import uuid
from decimal import Decimal

import structlog
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from oficinas.core.enums import StatusEntradaNfe, TipoMovimentacao
from oficinas.core.exceptions import EntradaJaProcessada, EstoqueInsuficiente, NaoEncontrado, NFeJaImportada
from oficinas.modules.estoque.models import (
    EntradaNfe,
    Fornecedor,
    ItemEntrada,
    MovimentacaoEstoque,
    Produto,
)
from oficinas.modules.estoque.parser import NFeParseResult, parse_nfe
from oficinas.modules.estoque.schemas import (
    FornecedorCreate,
    FornecedorUpdate,
    ProdutoCreate,
    ProdutoUpdate,
)

log = structlog.get_logger()

# Movimentações que aumentam o estoque físico disponível
_AUMENTAM = {TipoMovimentacao.ENTRADA, TipoMovimentacao.LIBERACAO}
# SAIDA é apenas registro contábil — o estoque já foi reduzido pela RESERVA
_SEM_EFEITO = {TipoMovimentacao.SAIDA}


class EstoqueService:
    def __init__(self, db: AsyncSession) -> None:
        self.db = db

    # ─── Fornecedor ───────────────────────────────────────────────────────────

    async def criar_fornecedor(self, tenant_id: uuid.UUID, payload: FornecedorCreate) -> Fornecedor:
        f = Fornecedor(tenant_id=tenant_id, **payload.model_dump())
        self.db.add(f)
        await self.db.commit()
        await self.db.refresh(f)
        log.info("fornecedor_criado", fornecedor_id=str(f.id))
        return f

    async def listar_fornecedores(
        self,
        tenant_id: uuid.UUID,
        q: str | None = None,
        ativo: bool | None = None,
        tipo_pessoa: str | None = None,
    ) -> list[Fornecedor]:
        stmt = select(Fornecedor).where(Fornecedor.tenant_id == tenant_id)
        if q:
            safe = q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            pattern = f"%{safe}%"
            from sqlalchemy import or_
            stmt = stmt.where(or_(
                Fornecedor.razao_social.ilike(pattern),
                Fornecedor.nome_fantasia.ilike(pattern),
                Fornecedor.cnpj.ilike(pattern),
            ))
        if ativo is not None:
            stmt = stmt.where(Fornecedor.ativo.is_(ativo))
        if tipo_pessoa:
            stmt = stmt.where(Fornecedor.tipo_pessoa == tipo_pessoa)
        stmt = stmt.order_by(Fornecedor.razao_social)
        return list((await self.db.execute(stmt)).scalars().all())

    async def importar_fornecedores_xlsx(
        self, tenant_id: uuid.UUID, conteudo: bytes
    ) -> dict:
        import io
        import openpyxl

        if conteudo[:4] == b"\xd0\xcf\x11\xe0":
            raise ValueError(
                "Formato .xls não suportado. Abra o arquivo no Excel e salve como "
                "'Pasta de Trabalho do Excel (.xlsx)' antes de importar."
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            raise ValueError(f"Arquivo inválido: {exc}") from exc

        if not rows:
            raise ValueError("Planilha vazia")

        cabecalhos = {str(c).strip().lower() if c else "": i for i, c in enumerate(rows[0])}

        def col(row: tuple, *nomes: str):
            for nome in nomes:
                idx = cabecalhos.get(nome.lower())
                if idx is not None and idx < len(row):
                    val = row[idx]
                    s = str(val).strip() if val is not None else None
                    return s if s else None
            return None

        criados = atualizados = ignorados = 0
        erros: list[str] = []

        for num, row in enumerate(rows[1:], start=2):
            try:
                razao = col(row, "razão social", "razao social", "nome")
                if not razao:
                    ignorados += 1
                    continue

                cnpj_raw = col(row, "cnpj", "cpf/cnpj", "cpf")
                cnpj = cnpj_raw.replace(".", "").replace("/", "").replace("-", "").replace(" ", "") if cnpj_raw else None
                if cnpj and len(cnpj) not in (11, 14):
                    cnpj = cnpj_raw  # mantém original se não normalizar

                status_str = col(row, "status") or "Ativo"
                ativo = status_str.lower() != "inativo"

                tipo_str = col(row, "tipo de pessoa", "tipo") or "Juridica"
                tipo_pessoa = "Fisica" if "fis" in tipo_str.lower() else "Juridica"

                dados = {
                    "razao_social":       razao,
                    "nome_fantasia":      col(row, "nome fantasia", "fantasia"),
                    "cnpj":               cnpj,
                    "inscricao_estadual": col(row, "inscrição estadual", "inscricao estadual", "ie"),
                    "telefone":           col(row, "telefone comercial", "telefone", "celular"),
                    "email":              col(row, "e-mail", "email"),
                    "contato":            col(row, "responsável", "responsavel", "contato"),
                    "ativo":              ativo,
                    "tipo_pessoa":        tipo_pessoa,
                }

                # Upsert por CNPJ dentro do tenant
                existente = None
                if cnpj:
                    stmt = select(Fornecedor).where(
                        Fornecedor.cnpj == cnpj,
                        Fornecedor.tenant_id == tenant_id,
                    )
                    existente = (await self.db.execute(stmt)).scalar_one_or_none()

                if existente:
                    for campo, valor in dados.items():
                        if valor is not None:
                            setattr(existente, campo, valor)
                    atualizados += 1
                else:
                    self.db.add(Fornecedor(tenant_id=tenant_id, **dados))
                    criados += 1

            except Exception as exc:
                erros.append(f"Linha {num}: {exc}")

        await self.db.commit()
        log.info("fornecedores_importados", criados=criados, atualizados=atualizados, tenant_id=str(tenant_id))
        return {"criados": criados, "atualizados": atualizados, "ignorados": ignorados, "erros": erros}

    async def listar_produtos_fornecedor(
        self, fornecedor_id: uuid.UUID, tenant_id: uuid.UUID, q: str | None = None
    ) -> list[dict]:
        from sqlalchemy import and_
        from oficinas.modules.estoque.models import MapeamentoFornecedorProduto
        stmt = (
            select(
                MapeamentoFornecedorProduto.id,
                MapeamentoFornecedorProduto.produto_id,
                MapeamentoFornecedorProduto.codigo_fornecedor,
                Produto.codigo,
                Produto.descricao,
                Produto.marca,
            )
            .join(Produto, MapeamentoFornecedorProduto.produto_id == Produto.id)
            .where(
                and_(
                    MapeamentoFornecedorProduto.fornecedor_id == fornecedor_id,
                    MapeamentoFornecedorProduto.tenant_id == tenant_id,
                )
            )
        )
        if q:
            safe = q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            pattern = f"%{safe}%"
            from sqlalchemy import or_
            stmt = stmt.where(or_(
                Produto.codigo.ilike(pattern),
                Produto.descricao.ilike(pattern),
                MapeamentoFornecedorProduto.codigo_fornecedor.ilike(pattern),
            ))
        stmt = stmt.order_by(Produto.descricao)
        rows = (await self.db.execute(stmt)).all()
        return [
            {
                "mapeamento_id": r[0],
                "produto_id":    r[1],
                "codigo_fornecedor": r[2],
                "codigo_interno": r[3],
                "descricao":     r[4],
                "marca":         r[5],
            }
            for r in rows
        ]

    async def buscar_fornecedor(self, fornecedor_id: uuid.UUID, tenant_id: uuid.UUID) -> Fornecedor:
        stmt = select(Fornecedor).where(Fornecedor.id == fornecedor_id, Fornecedor.tenant_id == tenant_id)
        f = (await self.db.execute(stmt)).scalar_one_or_none()
        if not f:
            raise NaoEncontrado(f"Fornecedor {fornecedor_id} não encontrado")
        return f

    async def atualizar_fornecedor(
        self, fornecedor_id: uuid.UUID, tenant_id: uuid.UUID, payload: FornecedorUpdate
    ) -> Fornecedor:
        f = await self.buscar_fornecedor(fornecedor_id, tenant_id)
        for campo, valor in payload.model_dump(exclude_unset=True).items():
            setattr(f, campo, valor)
        await self.db.commit()
        await self.db.refresh(f)
        return f

    # ─── Produto ──────────────────────────────────────────────────────────────

    async def importar_produtos_xlsx(
        self, tenant_id: uuid.UUID, conteudo: bytes
    ) -> dict:
        import io
        import openpyxl

        if conteudo[:4] == b"\xd0\xcf\x11\xe0":
            raise ValueError(
                "Formato .xls não suportado. Abra o arquivo no Excel e salve como "
                "'Pasta de Trabalho do Excel (.xlsx)' antes de importar."
            )

        try:
            wb = openpyxl.load_workbook(io.BytesIO(conteudo), data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
        except Exception as exc:
            raise ValueError(f"Arquivo inválido: {exc}") from exc

        if not rows:
            raise ValueError("Planilha vazia")

        cabecalhos = {str(c).strip().lower() if c else "": i for i, c in enumerate(rows[0])}

        def col(row: tuple, *nomes: str):
            for nome in nomes:
                idx = cabecalhos.get(nome.lower())
                if idx is not None and idx < len(row):
                    val = row[idx]
                    s = str(val).strip() if val is not None else None
                    return s if s else None
            return None

        # Pré-carrega todos os produtos do tenant para upsert eficiente (evita N+1)
        stmt = select(Produto).where(Produto.tenant_id == tenant_id)
        existentes = {p.codigo: p for p in (await self.db.execute(stmt)).scalars().all()}

        criados = atualizados = ignorados = 0
        erros: list[str] = []

        for num, row in enumerate(rows[1:], start=2):
            try:
                codigo = col(row, "código", "codigo")
                if not codigo:
                    ignorados += 1
                    continue

                descricao = col(row, "descrição", "descricao")
                if not descricao:
                    ignorados += 1
                    continue

                status_str = col(row, "status") or "Ativo"
                ativo = status_str.lower() != "inativo"

                marca = col(row, "marca")
                localizacao = col(row, "localização", "localizacao")

                valor_str = col(row, "valor")
                preco = Decimal(str(valor_str).replace(",", ".")) if valor_str else Decimal("0")

                if codigo in existentes:
                    p = existentes[codigo]
                    p.descricao = descricao
                    p.ativo = ativo
                    if marca is not None:
                        p.marca = marca
                    if localizacao is not None:
                        p.localizacao = localizacao
                    if preco > 0:
                        p.preco_venda = preco
                        p.preco_custo = preco
                    atualizados += 1
                else:
                    estoque_str = col(row, "estoque")
                    estoque = Decimal(str(estoque_str).replace(",", ".")) if estoque_str else Decimal("0")

                    p = Produto(
                        tenant_id=tenant_id,
                        codigo=codigo,
                        descricao=descricao,
                        marca=marca,
                        localizacao=localizacao,
                        preco_venda=preco,
                        preco_custo=preco,
                        estoque_atual=estoque,
                        ativo=ativo,
                    )
                    self.db.add(p)
                    existentes[codigo] = p
                    criados += 1

            except Exception as exc:
                erros.append(f"Linha {num}: {exc}")

        await self.db.commit()
        log.info("produtos_importados", criados=criados, atualizados=atualizados, tenant_id=str(tenant_id))
        return {"criados": criados, "atualizados": atualizados, "ignorados": ignorados, "erros": erros}

    async def criar_produto(self, tenant_id: uuid.UUID, payload: ProdutoCreate) -> Produto:
        p = Produto(tenant_id=tenant_id, estoque_atual=Decimal("0"), **payload.model_dump())
        self.db.add(p)
        await self.db.commit()
        await self.db.refresh(p)
        log.info("produto_criado", produto_id=str(p.id), codigo=p.codigo)
        return p

    async def listar_produtos(
        self,
        tenant_id: uuid.UUID,
        q: str | None = None,
        page: int = 1,
        page_size: int = 50,
    ) -> tuple[list[Produto], int]:
        from sqlalchemy import func, or_
        base = select(Produto).where(Produto.tenant_id == tenant_id, Produto.ativo.is_(True))
        if q:
            safe = q.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
            pattern = f"%{safe}%"
            base = base.where(or_(Produto.descricao.ilike(pattern), Produto.codigo.ilike(pattern)))

        total = (await self.db.execute(select(func.count()).select_from(base.subquery()))).scalar_one()

        offset = (page - 1) * page_size
        items = list(
            (await self.db.execute(base.order_by(Produto.descricao).offset(offset).limit(page_size)))
            .scalars()
            .all()
        )
        return items, total

    async def listar_marcas(self, tenant_id: uuid.UUID) -> list[str]:
        stmt = (
            select(Produto.marca)
            .where(Produto.tenant_id == tenant_id, Produto.ativo.is_(True), Produto.marca.isnot(None))
            .distinct()
            .order_by(Produto.marca)
        )
        return [r for r in (await self.db.execute(stmt)).scalars().all() if r]

    async def buscar_produto(self, produto_id: uuid.UUID, tenant_id: uuid.UUID) -> Produto:
        stmt = select(Produto).where(Produto.id == produto_id, Produto.tenant_id == tenant_id)
        p = (await self.db.execute(stmt)).scalar_one_or_none()
        if not p:
            raise NaoEncontrado(f"Produto {produto_id} não encontrado")
        return p

    async def atualizar_produto(
        self, produto_id: uuid.UUID, tenant_id: uuid.UUID, payload: ProdutoUpdate
    ) -> Produto:
        p = await self.buscar_produto(produto_id, tenant_id)
        for campo, valor in payload.model_dump(exclude_none=True).items():
            setattr(p, campo, valor)
        await self.db.commit()
        await self.db.refresh(p)
        return p

    # ─── Movimentação ─────────────────────────────────────────────────────────

    async def registrar_movimentacao(
        self,
        produto_id: uuid.UUID,
        tenant_id: uuid.UUID,
        tipo_mov: TipoMovimentacao,
        quantidade: Decimal,
        referencia_id: uuid.UUID | None = None,
        tipo_ref: str | None = None,
    ) -> MovimentacaoEstoque:
        """
        Registra uma movimentação e atualiza estoque_atual do produto.
        NÃO faz commit — o chamador é responsável pelo commit.

        ENTRADA / LIBERACAO  → estoque sobe
        RESERVA              → estoque desce (levanta EstoqueInsuficiente se saldo < 0)
        SAIDA                → apenas registro contábil; estoque não muda
                               (já foi reduzido pela RESERVA ao abrir o item na OS)
        """
        produto = await self.buscar_produto(produto_id, tenant_id)
        anterior = produto.estoque_atual

        if tipo_mov in _AUMENTAM:
            novo = anterior + quantidade
        elif tipo_mov in _SEM_EFEITO:
            novo = anterior
        else:  # RESERVA
            novo = anterior - quantidade
            if novo < 0:
                raise EstoqueInsuficiente(
                    f"Estoque insuficiente para '{produto.descricao}': "
                    f"disponível={anterior}, solicitado={quantidade}"
                )

        produto.estoque_atual = novo

        mov = MovimentacaoEstoque(
            tenant_id=tenant_id,
            produto_id=produto_id,
            referencia_id=referencia_id,
            tipo_ref=tipo_ref,
            tipo_mov=tipo_mov,
            quantidade=quantidade,
            estoque_anterior=anterior,
            estoque_novo=novo,
        )
        self.db.add(mov)
        log.info(
            "movimentacao_registrada",
            tipo=tipo_mov,
            produto_id=str(produto_id),
            anterior=str(anterior),
            novo=str(novo),
        )
        return mov

    async def listar_movimentacoes(
        self, produto_id: uuid.UUID, tenant_id: uuid.UUID
    ) -> list[MovimentacaoEstoque]:
        stmt = (
            select(MovimentacaoEstoque)
            .where(
                MovimentacaoEstoque.produto_id == produto_id,
                MovimentacaoEstoque.tenant_id == tenant_id,
            )
            .order_by(MovimentacaoEstoque.criado_em.desc())
        )
        return list((await self.db.execute(stmt)).scalars().all())

    # ─── NF-e de entrada ──────────────────────────────────────────────────────

    async def processar_entrada_xml(
        self, xml_bytes: bytes, tenant_id: uuid.UUID
    ) -> EntradaNfe:
        dados = parse_nfe(xml_bytes)  # levanta ValueError se XML inválido

        # Idempotência: chave única garante que a mesma NF-e não entre duas vezes
        if dados.chave:
            stmt = select(EntradaNfe).where(EntradaNfe.chave_nfe == dados.chave)
            if (await self.db.execute(stmt)).scalar_one_or_none():
                raise NFeJaImportada(f"NF-e {dados.chave} já importada")

        fornecedor = await self._upsert_fornecedor(tenant_id, dados)

        entrada = EntradaNfe(
            tenant_id=tenant_id,
            fornecedor_id=fornecedor.id if fornecedor else None,
            chave_nfe=dados.chave or None,
            numero_nf=dados.numero or None,
            data_emissao=dados.data_emissao or None,
            valor_total=dados.valor_total,
        )
        self.db.add(entrada)
        await self.db.flush()  # garante entrada.id antes de criar os itens

        for item_nfe in dados.itens:
            produto = await self._upsert_produto(tenant_id, item_nfe)

            self.db.add(ItemEntrada(
                entrada_id=entrada.id,
                produto_id=produto.id,
                codigo_fornecedor=item_nfe.codigo,
                quantidade=item_nfe.quantidade,
                preco_unitario=item_nfe.preco_unitario,
                icms=item_nfe.icms,
                ipi=item_nfe.ipi,
            ))

            await self.registrar_movimentacao(
                produto.id, tenant_id,
                TipoMovimentacao.ENTRADA, item_nfe.quantidade,
                referencia_id=entrada.id, tipo_ref="ENTRADA",
            )

        await self.db.commit()
        await self.db.refresh(entrada)
        log.info("nfe_importada", chave=dados.chave, tenant_id=str(tenant_id), itens=len(dados.itens))
        return entrada

    # ─── Helpers internos ─────────────────────────────────────────────────────

    async def _upsert_fornecedor(self, tenant_id: uuid.UUID, dados: NFeParseResult) -> Fornecedor | None:
        if not dados.emit_cnpj:
            return None
        stmt = select(Fornecedor).where(
            Fornecedor.cnpj == dados.emit_cnpj,
            Fornecedor.tenant_id == tenant_id,
        )
        f = (await self.db.execute(stmt)).scalar_one_or_none()
        if not f:
            f = Fornecedor(tenant_id=tenant_id, razao_social=dados.emit_nome, cnpj=dados.emit_cnpj)
            self.db.add(f)
            await self.db.flush()
        return f

    # ─── Entrada NF-e (pós-rascunho) ─────────────────────────────────────────

    async def buscar_entrada(self, entrada_id: uuid.UUID, tenant_id: uuid.UUID) -> EntradaNfe:
        stmt = select(EntradaNfe).where(
            EntradaNfe.id == entrada_id, EntradaNfe.tenant_id == tenant_id
        )
        e = (await self.db.execute(stmt)).scalar_one_or_none()
        if not e:
            raise NaoEncontrado(f"Entrada {entrada_id} não encontrada")
        return e

    async def listar_itens_entrada(self, entrada_id: uuid.UUID) -> list[ItemEntrada]:
        stmt = select(ItemEntrada).where(ItemEntrada.entrada_id == entrada_id)
        return list((await self.db.execute(stmt)).scalars().all())

    async def processar_entrada(
        self,
        entrada_id: uuid.UUID,
        tenant_id: uuid.UUID,
        data_entrada_nota,
        itens_update: list,
    ) -> EntradaNfe:
        entrada = await self.buscar_entrada(entrada_id, tenant_id)
        if entrada.status == StatusEntradaNfe.PROCESSADA:
            raise EntradaJaProcessada(f"Entrada {entrada_id} já está processada")

        if data_entrada_nota is not None:
            entrada.data_entrada = data_entrada_nota

        if itens_update:
            itens = await self.listar_itens_entrada(entrada_id)
            itens_por_id = {i.id: i for i in itens}
            for upd in itens_update:
                item = itens_por_id.get(upd.id)
                if item and upd.data_entrada is not None:
                    item.data_entrada = upd.data_entrada

        entrada.status = StatusEntradaNfe.PROCESSADA
        await self.db.commit()
        await self.db.refresh(entrada)
        log.info("entrada_processada", entrada_id=str(entrada_id))
        return entrada

    # ─── Helpers internos ─────────────────────────────────────────────────────

    async def _upsert_produto(self, tenant_id: uuid.UUID, item) -> Produto:
        stmt = select(Produto).where(
            Produto.codigo == item.codigo,
            Produto.tenant_id == tenant_id,
        )
        p = (await self.db.execute(stmt)).scalar_one_or_none()
        if p:
            p.preco_custo = item.preco_unitario  # atualiza preço de custo
            return p
        p = Produto(
            tenant_id=tenant_id,
            codigo=item.codigo,
            descricao=item.descricao,
            ncm=item.ncm or None,
            preco_custo=item.preco_unitario,
            preco_venda=item.preco_unitario,  # admin ajusta depois
            estoque_atual=Decimal("0"),
            ativo=True,
        )
        self.db.add(p)
        await self.db.flush()
        return p
